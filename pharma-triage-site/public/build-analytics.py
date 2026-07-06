# -*- coding: utf-8 -*-
"""
build-analytics.py  —  Pharma Triage call-analytics data builder.

Reads the two monthly call reports that live next to this script:
    - "june report customer .xlsx"   (calls from customers / patients)
    - "june report internal .xlsx"   (internal calls from colleagues / branches)

Each workbook has the same sheets (produced by the call-center exporter):
    Daily Summary | Shift Analysis | Daily Ranking | Employee Analysis |
    Employee Shift | Insights | Debug | Bad Dates

This script flattens what the dashboard needs and EMBEDS it directly into
`call-analytics.html` (between the  // DATA_START  /  // DATA_END  markers),
so the dashboard works fully offline with no server.

Re-run any time you drop in fresh files (keep the same names, or edit FILES):

    python build-analytics.py

Requires: openpyxl   (pip install openpyxl)
"""
import os, re, json, datetime as dt
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_JSON = os.path.join(HERE, "analytics-data.json")
HTML_FILE = os.path.join(HERE, "call-analytics.html")

# ── Monthly file registry ────────────────────────────────────────────────────
# Each key is the Arabic month label shown in the dashboard selector.
# To add a new month:
#   1. Drop the two .xlsx files next to this script.
#   2. Uncomment (or add) the matching entry below.
#   3. Run:  python build-analytics.py
# The dashboard will automatically show a month-selector tab when >1 month exists.
# ─────────────────────────────────────────────────────────────────────────────
MONTHLY_FILES = {
    "يونيو ٢٠٢٦": {
        "customer": "june report customer .xlsx",   # استشارات دوائية — عملاء (110)
        "internal": "june report internal .xlsx",    # استشارات دوائية — موظفين (111)
        "cosmetic": "cosmo.xlsx",                    # استشارات تجميلية (107) — حُط الملف هنا وشغّل السكربت
    },
    # "يوليو ٢٠٢٦": {
    #     "customer": "july report customer .xlsx",
    #     "internal": "july report internal .xlsx",
    # },
    # "أغسطس ٢٠٢٦": {
    #     "customer": "august report customer .xlsx",
    #     "internal": "august report internal .xlsx",
    # },
}

# Raw shift label -> short Arabic label
SHIFTS = {
    "12.00 AM : 7.00 AM": "ليلي · ١٢ص–٧ص",
    "7:00 AM : 1.00 PM":  "صباحي · ٧ص–١م",
    "1.00 PM : 5.00 PM":  "ظهري · ١م–٥م",
    "5.00 PM : 12.00 AM": "مسائي · ٥م–١٢ص",
}
SHIFT_ORDER = list(SHIFTS.keys())


def iso(d):
    """'21/04/2026' -> '2026-04-21'."""
    if isinstance(d, dt.datetime):
        return d.date().isoformat()
    m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})", str(d))
    if not m:
        return None
    dd, mm, yy = m.groups()
    return f"{yy}-{int(mm):02d}-{int(dd):02d}"


def num(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return round(v, 4)
    return v


def sheet(wb, name):
    for ws in wb.worksheets:
        if ws.title.strip().lower() == name.lower():
            return ws
    return None


def col(ws, r, c):
    return ws.cell(row=r, column=c).value


def parse_daily(ws):
    """Daily Summary -> chronological per-day totals (from the 'TOTAL =' rows)
    plus the GRAND TOTAL row."""
    days, grand, cur = [], None, None
    for r in range(2, ws.max_row + 1):
        a, b = col(ws, r, 1), col(ws, r, 2)
        if a not in (None, ""):
            cur = iso(a)
        bs = str(b).strip() if b is not None else ""
        if bs == "TOTAL =":
            days.append({
                "date": cur,
                "total": num(col(ws, r, 3)) or 0,
                "abandoned": num(col(ws, r, 4)) or 0,
                "answered": num(col(ws, r, 5)) or 0,
                "rate": num(col(ws, r, 6)) or 0,
                "status": col(ws, r, 7),
            })
        elif bs == "GRAND TOTAL =":
            grand = {
                "total": num(col(ws, r, 3)) or 0,
                "abandoned": num(col(ws, r, 4)) or 0,
                "answered": num(col(ws, r, 5)) or 0,
                "rate": num(col(ws, r, 6)) or 0,
            }
    # drop trailing all-zero days (e.g. the cut-off last day)
    while days and days[-1]["total"] == 0:
        days.pop()
    return days, grand


def parse_shifts(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        b = col(ws, r, 2)
        if b in (None, ""):
            continue
        out.append({
            "raw": str(b).strip(),
            "shift": SHIFTS.get(str(b).strip(), str(b).strip()),
            "total": num(col(ws, r, 3)) or 0,
            "abandoned": num(col(ws, r, 4)) or 0,
            "answered": num(col(ws, r, 5)) or 0,
            "rate": num(col(ws, r, 6)) or 0,
            "avg_day": num(col(ws, r, 7)) or 0,
            "days": num(col(ws, r, 9)) or 0,
            "status": col(ws, r, 10),
        })
    out.sort(key=lambda s: SHIFT_ORDER.index(s["raw"]) if s["raw"] in SHIFT_ORDER else 99)
    return out


def parse_employees(ws):
    out = []
    for r in range(2, ws.max_row + 1):
        name = col(ws, r, 2)
        if name in (None, ""):
            continue
        out.append({
            "name": str(name).strip(),
            "answered": num(col(ws, r, 3)) or 0,
            "pct": num(col(ws, r, 4)) or 0,
        })
    out.sort(key=lambda e: e["answered"], reverse=True)
    return out


def parse_insights(ws):
    d = {}
    if not ws:
        return d
    for r in range(2, ws.max_row + 1):
        k, v = col(ws, r, 1), col(ws, r, 2)
        if k:
            d[str(k).strip()] = num(v)
    return d


def parse_consult_team(ws):
    """
    CONSULTATION TEAM ANALYSIS sheet parser.

    Expected columns (row 1 = headers, data from row 2):
        A  رقم الدخول   – agent login number  (e.g. 101, 173 …)
        B  الاسم        – display name  (same name used in the call system)
        C  عدد الاستشارات – total consultations handled this month

    Example rows:
        101 | عمر الازهري فوزي عبدالحميد    | 245
        102 | باسم نبيل صبحي ميخائيل        | 189
        173 | دينا محمد يحيى محمد شطا        | 134

    Tips for next month:
    - Keep the same three columns — the script auto-detects by position.
    - Column B (الاسم) should ideally match how the agent appears in the
      call-center "Employee Analysis" sheet.  The dashboard uses a smart
      Arabic normalizer (handles alef variants, ta-marbuta, extra spaces,
      harakat) so minor spelling differences are tolerated.
    - If a match still fails, check for big differences (e.g. nickname vs
      full name).  You can open the browser console → look for "unmatched"
      log messages to debug.
    - Sort order doesn't matter — the script re-sorts by internal calls desc.
    """
    if not ws:
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        login_id = col(ws, r, 1)
        name     = col(ws, r, 2)
        consults = col(ws, r, 3)
        if login_id is None and name is None:
            continue
        out.append({
            "login_id":    str(int(login_id)).strip() if isinstance(login_id, (int, float)) else str(login_id or "").strip(),
            "name":        str(name or "").strip(),
            "consultations": int(num(consults) or 0),
        })
    out.sort(key=lambda e: e["consultations"], reverse=True)
    return [e for e in out if e["consultations"] > 0 or e["name"]]


def parse_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    daily, grand = parse_daily(sheet(wb, "Daily Summary"))
    shifts = parse_shifts(sheet(wb, "Shift Analysis"))
    employees = parse_employees(sheet(wb, "Employee Analysis"))
    insights = parse_insights(sheet(wb, "Insights"))
    consult_team = parse_consult_team(sheet(wb, "CONSULTATION TEAM ANALYSIS"))
    wb.close()

    period = None
    if daily:
        period = f"{daily[0]['date']} → {daily[-1]['date']}"

    result = {
        "period": period,
        "totals": grand or {},
        "days": daily,
        "shifts": shifts,
        "employees": employees,
        "n_employees": len(employees),
        "insights": insights,
    }
    if consult_team:
        result["consult_team"] = consult_team
    return result


def main():
    month_keys = list(MONTHLY_FILES.keys())
    if not month_keys:
        raise SystemExit("MONTHLY_FILES is empty — add at least one month entry.")

    out = {
        "generated": dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "months": {},
        "datasets": {},   # backward-compat: points to latest month
    }

    for month_label, files in MONTHLY_FILES.items():
        month_data = {}
        for key, fname in files.items():
            path = os.path.join(HERE, fname)
            if not os.path.exists(path):
                # ملف اختياري (زي التجميلية) لسه ماتحطّش — يتخطّى بتحذير بدل ما يكسر البناء
                print(f"  ⏭️  [{month_label}] تخطّي '{key}' — الملف غير موجود: {fname}")
                continue
            month_data[key] = parse_file(path)
        if not month_data:
            raise SystemExit(f"Missing ALL files for [{month_label}] — حُط ملف واحد على الأقل.")
        out["months"][month_label] = month_data
        print(f"\n  ── {month_label} ──")
        for k, d in month_data.items():
            t = d["totals"]
            print(f"    {k:9s}: total={t.get('total')}  answered={t.get('answered')}  "
                  f"abandoned={t.get('abandoned')}  days={len(d['days'])}  "
                  f"staff={d['n_employees']}")

    # datasets = latest month (backward compat for old dashboard versions)
    out["datasets"] = out["months"][month_keys[-1]]

    payload = json.dumps(out, ensure_ascii=False, separators=(",", ":"))
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        f.write(payload)

    if os.path.exists(HTML_FILE):
        html = open(HTML_FILE, encoding="utf-8").read()
        block = "// DATA_START\nconst DASH_DATA = " + payload + ";\n// DATA_END"
        new = re.sub(r"// DATA_START.*?// DATA_END", lambda _: block, html, flags=re.S)
        if new != html:
            open(HTML_FILE, "w", encoding="utf-8").write(new)
            print(f"\nEmbedded {len(month_keys)} month(s) into call-analytics.html")
        else:
            print("WARNING: // DATA_START .. // DATA_END markers not found.")

    print(f"\nwrote analytics-data.json ({len(payload)} bytes)  |  months: {', '.join(month_keys)}")


if __name__ == "__main__":
    main()
